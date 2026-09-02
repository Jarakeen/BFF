from __future__ import annotations

"""Canonical, round-trip-friendly achievement progress exports."""

import csv
from pathlib import Path

from services.achievement_progress_service import AchievementProgressService
from services.eso_achievement_database_service import EsoAchievementDatabaseService


class AchievementProgressExportService:
    SCHEMA_VERSION = 1

    def __init__(
        self,
        achievement_data: EsoAchievementDatabaseService,
        achievement_progress: AchievementProgressService,
    ) -> None:
        self.achievement_data = achievement_data
        self.achievement_progress = achievement_progress

    def _achievement_rows(self):
        return self.achievement_data.connection.execute(
            """
            SELECT a.id, a.name,
                   COALESCE(c.category_name, '') AS category_name,
                   COALESCE(c.subcategory_name, '') AS subcategory_name,
                   COALESCE(a.points, 0) AS points,
                   a.collectible_id
            FROM achievement a
            LEFT JOIN achievement_category c
              ON c.category_index = a.category_index
             AND c.subcategory_index = a.subcategory_index
            ORDER BY c.category_index, c.subcategory_index, a.achievement_index, a.id
            """
        ).fetchall()

    def export_csv(self, target_path: Path) -> Path:
        """Write one normalized row per achievement/profile pair."""
        target_path = Path(target_path)
        target_path.parent.mkdir(parents=True, exist_ok=True)
        profiles = self.achievement_progress.profiles()
        rows = self._achievement_rows()

        with target_path.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.writer(handle)
            writer.writerow(
                [
                    "Schema Version",
                    "Achievement ID",
                    "Name",
                    "Category",
                    "Subcategory",
                    "Points",
                    "Collectible Reward ID",
                    "Profile",
                    "Completed",
                ]
            )
            for row in rows:
                achievement_id = str(row["id"])
                for profile in profiles:
                    writer.writerow(
                        [
                            self.SCHEMA_VERSION,
                            achievement_id,
                            row["name"],
                            row["category_name"],
                            row["subcategory_name"],
                            row["points"],
                            row["collectible_id"] if row["collectible_id"] is not None else "",
                            profile,
                            1 if achievement_id in self.achievement_progress.completed_ids(profile) else 0,
                        ]
                    )
        return target_path

    def export_xlsx(self, target_path: Path) -> Path:
        """Write a normalized workbook with Achievements, Progress, and Meta sheets."""
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise RuntimeError(
                "Spreadsheet export needs openpyxl. Install it with: pip install openpyxl"
            ) from exc

        target_path = Path(target_path)
        target_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        achievements = workbook.active
        achievements.title = "Achievements"
        achievements.append(
            [
                "Achievement ID",
                "Name",
                "Category",
                "Subcategory",
                "Points",
                "Collectible Reward ID",
            ]
        )
        rows = self._achievement_rows()
        for row in rows:
            achievements.append(
                [
                    int(row["id"]),
                    row["name"],
                    row["category_name"],
                    row["subcategory_name"],
                    int(row["points"] or 0),
                    int(row["collectible_id"]) if row["collectible_id"] is not None else None,
                ]
            )

        progress = workbook.create_sheet("Progress")
        progress.append(["Profile", "Achievement ID", "Completed"])
        profiles = self.achievement_progress.profiles()
        for profile in profiles:
            completed = self.achievement_progress.completed_ids(profile)
            for row in rows:
                achievement_id = str(row["id"])
                progress.append([profile, int(row["id"]), 1 if achievement_id in completed else 0])

        meta = workbook.create_sheet("Meta")
        meta.append(["Key", "Value"])
        meta.append(["Schema Version", self.SCHEMA_VERSION])
        meta.append(["Format", "Foundry Dock Achievement Progress"])
        meta.append(["Profiles", ", ".join(profiles)])

        workbook.save(target_path)
        workbook.close()
        return target_path
