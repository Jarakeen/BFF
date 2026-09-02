from __future__ import annotations

"""Read legacy and Foundry-native achievement workbooks without Google APIs."""

from pathlib import Path

from services.google_sheets_service import (
    ACHIEVEMENT_TABS,
    COLUMN_FOR_PERSON,
    GoogleSheetAchievementSnapshot,
    GoogleSheetAchievementStatus,
    NAME_COL,
    POINTS_COL,
)


class LocalAchievementWorkbookService:
    """Read .xlsx/.xlsm achievement progress exported from Google Sheets.

    Legacy BFF workbooks use column A for Rylo, B for Jarakeen, C for the
    achievement name, and F for points. New Foundry-native workbooks use
    canonical achievement IDs in separate ``Achievements`` and ``Progress``
    sheets.
    """

    @staticmethod
    def _is_checked(value) -> bool:
        return str(value or "").strip().casefold() in {"x", "✓", "✔", "yes", "true", "1"}

    @staticmethod
    def _load_workbook(path: Path):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "Local spreadsheet import needs openpyxl. Install it with: pip install openpyxl"
            ) from exc

        path = Path(path)
        if path.suffix.casefold() not in {".xlsx", ".xlsm"}:
            raise ValueError("Choose an .xlsx or .xlsm workbook.")
        if not path.exists():
            raise FileNotFoundError(path)
        return load_workbook(path, read_only=True, data_only=True)

    def read_person(self, path: Path, person: str) -> GoogleSheetAchievementSnapshot:
        person = str(person or "").strip()
        if person not in COLUMN_FOR_PERSON:
            raise ValueError(f"Unknown profile source: {person}")

        workbook = self._load_workbook(path)
        try:
            if "Achievements" in workbook.sheetnames and "Progress" in workbook.sheetnames:
                return self._read_foundry_native(workbook, person)
            return self._read_legacy(workbook, person)
        finally:
            workbook.close()

    def _read_legacy(self, workbook, person: str) -> GoogleSheetAchievementSnapshot:
        col = COLUMN_FOR_PERSON[person]
        rows: list[GoogleSheetAchievementStatus] = []
        missing_tabs: list[str] = []

        for tab_name in ACHIEVEMENT_TABS:
            if tab_name not in workbook.sheetnames:
                missing_tabs.append(tab_name)
                continue
            sheet = workbook[tab_name]
            for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
                name = str(values[NAME_COL - 1] or "").strip() if len(values) >= NAME_COL else ""
                points = values[POINTS_COL - 1] if len(values) >= POINTS_COL else None
                if not name or points in (None, ""):
                    continue
                mark = values[col - 1] if len(values) >= col else ""
                rows.append(
                    GoogleSheetAchievementStatus(
                        tab_name=tab_name,
                        row_number=row_number,
                        name=name,
                        checked=self._is_checked(mark),
                    )
                )

        return GoogleSheetAchievementSnapshot(
            person=person,
            rows=tuple(rows),
            missing_tabs=tuple(missing_tabs),
        )

    def _read_foundry_native(self, workbook, person: str) -> GoogleSheetAchievementSnapshot:
        achievement_names: dict[str, str] = {}
        achievements = workbook["Achievements"]
        header = [str(value or "").strip() for value in next(achievements.iter_rows(values_only=True), ())]
        header_map = {name.casefold(): index for index, name in enumerate(header)}
        id_index = header_map.get("achievement id")
        name_index = header_map.get("name")
        if id_index is None or name_index is None:
            raise ValueError("Foundry workbook Achievements sheet is missing Achievement ID or Name columns.")

        for values in achievements.iter_rows(values_only=True):
            if len(values) <= max(id_index, name_index):
                continue
            achievement_id = str(values[id_index] or "").strip()
            name = str(values[name_index] or "").strip()
            if achievement_id and name and achievement_id.casefold() != "achievement id":
                achievement_names[achievement_id] = name

        progress = workbook["Progress"]
        progress_header = [str(value or "").strip() for value in next(progress.iter_rows(values_only=True), ())]
        progress_map = {name.casefold(): index for index, name in enumerate(progress_header)}
        profile_index = progress_map.get("profile")
        id_index = progress_map.get("achievement id")
        completed_index = progress_map.get("completed")
        if None in (profile_index, id_index, completed_index):
            raise ValueError("Foundry workbook Progress sheet is missing Profile, Achievement ID, or Completed columns.")

        rows: list[GoogleSheetAchievementStatus] = []
        for row_number, values in enumerate(progress.iter_rows(values_only=True), start=1):
            if len(values) <= max(profile_index, id_index, completed_index):
                continue
            profile = str(values[profile_index] or "").strip()
            if profile.casefold() != person.casefold():
                continue
            achievement_id = str(values[id_index] or "").strip()
            name = achievement_names.get(achievement_id, "")
            if not name:
                continue
            rows.append(
                GoogleSheetAchievementStatus(
                    tab_name="Progress",
                    row_number=row_number,
                    name=name,
                    checked=self._is_checked(values[completed_index]),
                )
            )

        return GoogleSheetAchievementSnapshot(person=person, rows=tuple(rows), missing_tabs=())
