from __future__ import annotations

"""Read standalone collectible ownership from legacy BFF workbooks."""

from dataclasses import dataclass
from pathlib import Path

from services.google_sheets_service import ACHIEVEMENT_TABS, COLUMN_FOR_PERSON


@dataclass(frozen=True)
class LocalCollectibleRow:
    tab_name: str
    row_number: int
    name: str
    checked: bool


@dataclass(frozen=True)
class LocalCollectibleSnapshot:
    person: str
    rows: tuple[LocalCollectibleRow, ...]
    scanned_tabs: tuple[str, ...]

    @property
    def checked_rows(self) -> tuple[LocalCollectibleRow, ...]:
        return tuple(row for row in self.rows if row.checked)


class LocalCollectibleWorkbookService:
    """Extract collectible checkmarks without treating the workbook as canonical data.

    Legacy workbooks were allowed to evolve organically, so this reader first
    looks for explicit headers (Jarakeen/Rylo plus Name/Collectible) and only
    falls back to the old A/B/C convention on sheets whose names look like
    collectible categories. Achievement tabs are always excluded.
    """

    _NAME_HEADERS = {"name", "collectible", "collectible name", "item", "item name"}
    _SHEET_HINTS = (
        "mount", "pet", "assistant", "ally", "allies", "house", "costume", "skin",
        "polymorph", "personality", "hairstyle", "adornment", "memento", "emote",
        "customized action", "weapon style", "armor style", "furnishing", "fragment",
        "tool", "upgrade", "collectible", "collection",
    )

    @staticmethod
    def _is_checked(value) -> bool:
        return str(value or "").strip().casefold() in {"x", "✓", "✔", "yes", "true", "1"}

    @staticmethod
    def _load_workbook(path: Path):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "Local spreadsheet import needs openpyxl. Install it with: pip install openpyxl"
            ) from exc

        path = Path(path)
        if path.suffix.casefold() not in {".xlsx", ".xlsm"}:
            raise ValueError("Choose an .xlsx or .xlsm workbook.")
        if not path.exists():
            raise FileNotFoundError(path)
        return load_workbook(path, read_only=True, data_only=True)

    @classmethod
    def _looks_collectible_sheet(cls, title: str) -> bool:
        value = " ".join(str(title or "").casefold().replace("_", " ").split())
        return any(hint in value for hint in cls._SHEET_HINTS)

    @classmethod
    def _find_header(cls, sheet, person: str) -> tuple[int, int, int] | None:
        person_key = person.casefold()
        for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_number > 20:
                break
            normalized = [" ".join(str(value or "").strip().casefold().split()) for value in values]
            person_index = next((i for i, value in enumerate(normalized) if value == person_key), None)
            name_index = next((i for i, value in enumerate(normalized) if value in cls._NAME_HEADERS), None)
            if person_index is not None and name_index is not None:
                return row_number, person_index, name_index
        return None

    def read_person(self, path: Path, person: str) -> LocalCollectibleSnapshot:
        person = str(person or "").strip()
        if person not in COLUMN_FOR_PERSON:
            raise ValueError(f"Unknown profile source: {person}")

        workbook = self._load_workbook(path)
        try:
            rows: list[LocalCollectibleRow] = []
            scanned_tabs: list[str] = []
            achievement_tabs = {name.casefold() for name in ACHIEVEMENT_TABS}
            reserved = {"achievements", "progress", "meta"}

            for tab_name in workbook.sheetnames:
                key = tab_name.casefold()
                if key in achievement_tabs or key in reserved:
                    continue
                sheet = workbook[tab_name]
                header = self._find_header(sheet, person)
                if header is None and not self._looks_collectible_sheet(tab_name):
                    continue

                scanned_tabs.append(tab_name)
                if header is not None:
                    header_row, person_index, name_index = header
                    for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
                        if row_number <= header_row or len(values) <= max(person_index, name_index):
                            continue
                        name = str(values[name_index] or "").strip()
                        if not name:
                            continue
                        rows.append(
                            LocalCollectibleRow(
                                tab_name=tab_name,
                                row_number=row_number,
                                name=name,
                                checked=self._is_checked(values[person_index]),
                            )
                        )
                    continue

                # Legacy fallback: A=Rylo, B=Jarakeen, C=name.
                person_index = COLUMN_FOR_PERSON[person] - 1
                name_index = 2
                for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if len(values) <= max(person_index, name_index):
                        continue
                    name = str(values[name_index] or "").strip()
                    if not name or name.casefold() in self._NAME_HEADERS:
                        continue
                    rows.append(
                        LocalCollectibleRow(
                            tab_name=tab_name,
                            row_number=row_number,
                            name=name,
                            checked=self._is_checked(values[person_index]),
                        )
                    )

            return LocalCollectibleSnapshot(
                person=person,
                rows=tuple(rows),
                scanned_tabs=tuple(scanned_tabs),
            )
        finally:
            workbook.close()
