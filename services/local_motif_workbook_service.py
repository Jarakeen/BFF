from __future__ import annotations

"""Read profile-aware motif completion from the legacy BFF workbook."""

from dataclasses import dataclass
from pathlib import Path


@dataclass(frozen=True)
class LocalMotifRow:
    tab_name: str
    row_number: int
    motif_number: int
    style_name: str
    checked: bool


@dataclass(frozen=True)
class LocalMotifSnapshot:
    person: str
    rows: tuple[LocalMotifRow, ...]
    source_tabs: tuple[str, ...]

    @property
    def checked_rows(self) -> tuple[LocalMotifRow, ...]:
        return tuple(row for row in self.rows if row.checked)


class LocalMotifWorkbookService:
    """Extract completed motif rows from the old Jarakeen/Rylo motif sheets.

    The historical workbook stores motif completion separately from ordinary
    collectibles. Both motif sheets use B=R, C=J, D=motif number, E=style name.
    The chapter columns are retained in the workbook for partial tracking, but
    the supplied workbook currently records ownership at the whole-motif row.
    """

    PERSON_COLUMN = {"Rylo": 2, "Jarakeen": 3}  # one-based Excel columns
    SOURCE_TABS = ("Rylos Motifs", "Jarakeens Motifs")

    @staticmethod
    def _is_checked(value) -> bool:
        return str(value or "").strip().casefold() in {"x", "✓", "✔", "yes", "true", "1"}

    @staticmethod
    def _load_workbook(path: Path):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "Local spreadsheet import needs openpyxl. Install it with: pip install openpyxl"
            ) from exc

        path = Path(path)
        if path.suffix.casefold() not in {".xlsx", ".xlsm"}:
            raise ValueError("Choose an .xlsx or .xlsm workbook.")
        if not path.exists():
            raise FileNotFoundError(path)
        return load_workbook(path, read_only=True, data_only=True)

    def read_person(self, path: Path, person: str) -> LocalMotifSnapshot:
        person = str(person or "").strip()
        if person not in self.PERSON_COLUMN:
            raise ValueError(f"Unknown motif profile source: {person}")

        workbook = self._load_workbook(path)
        try:
            rows: list[LocalMotifRow] = []
            source_tabs: list[str] = []
            mark_index = self.PERSON_COLUMN[person] - 1

            for tab_name in self.SOURCE_TABS:
                if tab_name not in workbook.sheetnames:
                    continue
                source_tabs.append(tab_name)
                sheet = workbook[tab_name]
                for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if len(values) < 5:
                        continue
                    try:
                        motif_number = int(values[3])
                    except (TypeError, ValueError):
                        continue
                    style_name = " ".join(str(values[4] or "").strip().split())
                    if not style_name:
                        continue
                    mark = values[mark_index] if len(values) > mark_index else None
                    rows.append(
                        LocalMotifRow(
                            tab_name=tab_name,
                            row_number=row_number,
                            motif_number=motif_number,
                            style_name=style_name,
                            checked=self._is_checked(mark),
                        )
                    )

            return LocalMotifSnapshot(
                person=person,
                rows=tuple(rows),
                source_tabs=tuple(source_tabs),
            )
        finally:
            workbook.close()
